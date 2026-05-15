import os
import glob
import csv
import argparse
import numpy as np
from scipy.optimize import minimize
from openpyxl import load_workbook

R = 1.987e-3
TEMP = 297.0
INVALID_PROTEIN_PENALTY = 1e6

AA_ORDER = [
    "ALA", "ARG", "ASN", "ASP", "CYS",
    "GLN", "GLU", "GLY", "HIS", "ILE",
    "LEU", "LYS", "MET", "PHE", "PRO",
    "SER", "THR", "TRP", "TYR", "VAL",
]
AA_TO_IDX = {aa: i for i, aa in enumerate(AA_ORDER)}

MAX_SASA = {
    "ALA": 129.0, "ARG": 274.0, "ASN": 195.0, "ASP": 193.0, "CYS": 167.0,
    "GLN": 225.0, "GLU": 223.0, "GLY": 104.0, "HIS": 224.0, "ILE": 197.0,
    "LEU": 201.0, "LYS": 236.0, "MET": 224.0, "PHE": 240.0, "PRO": 159.0,
    "SER": 155.0, "THR": 172.0, "TRP": 285.0, "TYR": 263.0, "VAL": 174.0,
}

call_counter = {"n": 0}
best_tracker = {"best_loss": np.inf, "best_params": None, "invalid_ids": [], "invalid_details": []}


# parsing helpers
def normalize_text(s):
    return "" if s is None else str(s).replace("\xa0", " ").strip()


def to_float_safe(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = normalize_text(x).replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def extract_protein_id_from_npz_path(npz_path):
    name = os.path.splitext(os.path.basename(npz_path))[0]
    return name[:-5] if name.endswith("_full") else name


# remove local contacts from the contact map
def remove_local_contacts(contact_matrix, min_seq_sep=3):
    C = contact_matrix.copy()
    N = C.shape[0]
    for i in range(N):
        for j in range(N):
            if abs(i - j) < min_seq_sep:
                C[i, j] = 0
    return C


# prefix sums make block energy queries fast
def build_2d_prefix_sum(matrix):
    prefix = np.zeros((matrix.shape[0] + 1, matrix.shape[1] + 1), dtype=np.float64)
    prefix[1:, 1:] = np.cumsum(np.cumsum(matrix, axis=0), axis=1)
    return prefix


def rect_sum(prefix, r0, r1, c0, c1):
    return (
        prefix[r1 + 1, c1 + 1]
        - prefix[r0, c1 + 1]
        - prefix[r1 + 1, c0]
        + prefix[r0, c0]
    )


# read experimental folding dG targets from Excel
def load_excel_targets(excel_path, sheet_name=None):
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0] if sheet_name is None else sheet_name]

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise ValueError("Excel file is empty")

    header_norm = [normalize_text(h) for h in header]
    pdb_col_idx = None
    dg_col_idx = None

    for i, h in enumerate(header_norm):
        h_low = h.lower()
        if h_low in {"pdbid", "pbdid", "pdb id", "pbd id"}:
            pdb_col_idx = i
        if "folding dg" in h_low and "kcal/mol" in h_low:
            dg_col_idx = i

    if pdb_col_idx is None:
        raise ValueError(f"pdbID/pbdID column not found. Columns: {header_norm}")
    if dg_col_idx is None:
        raise ValueError(f"folding dG kcal/mol column not found. Columns: {header_norm}")

    targets = {}
    for row in rows:
        row_id = normalize_text(row[pdb_col_idx]) if pdb_col_idx < len(row) else ""
        dg_value = to_float_safe(row[dg_col_idx] if dg_col_idx < len(row) else None)
        if row_id and dg_value is not None:
            targets[row_id] = dg_value
    return targets


# Initial guess for per-amino-acid entropy parameters
def get_initial_entropy_params_per_aa(default_value=-0.01):
    return np.full(len(AA_ORDER), default_value, dtype=np.float64)


def get_sasa_path_for_protein(sasa_dir, protein_id):
    return os.path.join(sasa_dir, f"{protein_id}_residue_sasa.npy")


# Normalize SASA by the maximal area of the matching amino acid
def normalize_sasa_by_aa(res_name3, sasa_array, clip_min=0.0, clip_max=1.5):
    if len(res_name3) != len(sasa_array):
        raise ValueError(f"Length mismatch: res_name3={len(res_name3)}, sasa_array={len(sasa_array)}")

    sasa_norm = np.zeros(len(res_name3), dtype=np.float64)
    for i, aa in enumerate(res_name3):
        aa = str(aa).strip().upper()
        if aa not in MAX_SASA:
            raise KeyError(f"Unknown amino acid for SASA normalization: {aa}")
        sasa_norm[i] = sasa_array[i] / MAX_SASA[aa]

    return np.clip(sasa_norm, clip_min, clip_max)


# Build the entropy vector using one fitted parameter per amino acid
def build_entropy_vector_from_sasa_per_aa(res_name3, sasa_norm, entropy_aa_params):
    if len(res_name3) != len(sasa_norm):
        raise ValueError(f"Length mismatch: res_name3={len(res_name3)}, sasa_norm={len(sasa_norm)}")
    if len(entropy_aa_params) != len(AA_ORDER):
        raise ValueError(f"Expected {len(AA_ORDER)} entropy params, got {len(entropy_aa_params)}")

    entropy_vector = np.zeros(len(res_name3), dtype=np.float64)
    for i, aa in enumerate(res_name3):
        aa = str(aa).strip().upper()
        if aa not in AA_TO_IDX:
            raise KeyError(f"Unknown amino acid in res_name3: {aa}")
        burial_factor = 1.0 - sasa_norm[i]
        entropy_vector[i] = entropy_aa_params[AA_TO_IDX[aa]] * burial_factor
    return entropy_vector


# load one protein and prepare matrices needed by the model
def load_single_protein_inputs(npz_path, sasa_path, cutoff="C8", min_seq_sep=3):
    data = np.load(npz_path, allow_pickle=True)
    if "T" not in data:
        raise KeyError("File has no value 'T'")
    if cutoff not in data:
        raise KeyError(f"File has no value '{cutoff}'")
    if "res_name3" not in data:
        raise KeyError("File has no value 'res_name3'")

    T_tensor = data["T"].astype(np.float64)
    contact_matrix = data[cutoff].astype(np.float64)
    res_name3 = np.array([str(x).strip().upper() for x in data["res_name3"]], dtype=object)
    sasa_array = np.load(sasa_path).astype(np.float64)

    N = T_tensor.shape[0]
    if T_tensor.shape != (N, N, 5):
        raise ValueError(f"Expected T shape = (N, N, 5), actual T shape = {T_tensor.shape}")
    if contact_matrix.shape != (N, N):
        raise ValueError(f"Size contact matrix {contact_matrix.shape} doesnt match T {T_tensor.shape}")
    if len(res_name3) != N:
        raise ValueError(f"Length res_name3 ({len(res_name3)}) doesnt match T ({N})")
    if len(sasa_array) != N:
        raise ValueError(f"Length sasa_array ({len(sasa_array)}) doesnt match T ({N})")

    # Take raw VDW and electrostatic matrices from T tensor
    vdw_matrix = T_tensor[:, :, 0].copy()
    elec_matrix = T_tensor[:, :, 2].copy()

    # Normalize VDW and electrostatic matrices by their maximum absolute values.
    # This makes both energy terms comparable in scale before fitting.
    vdw_max_abs = np.max(np.abs(vdw_matrix))
    elec_max_abs = np.max(np.abs(elec_matrix))

    if vdw_max_abs > 0.0 and np.isfinite(vdw_max_abs):
        vdw_matrix = vdw_matrix / vdw_max_abs

    if elec_max_abs > 0.0 and np.isfinite(elec_max_abs):
        elec_matrix = elec_matrix / elec_max_abs

    sasa_norm = normalize_sasa_by_aa(res_name3, sasa_array)

    return {
        "contact_matrix": remove_local_contacts(contact_matrix, min_seq_sep=min_seq_sep),
        "vdw_matrix": vdw_matrix,
        "elec_matrix": elec_matrix,
        "res_name3": res_name3,
        "sasa_norm": sasa_norm,
    }


# Optionally restrict the fit to a subset of proteins
def select_dataset_subset(dataset, protein_limit=None, protein_ids=None):
    if protein_ids:
        wanted = {normalize_text(x) for x in protein_ids if normalize_text(x)}
        dataset = [item for item in dataset if item["protein_id"] in wanted]

    if protein_limit is not None:
        if protein_limit <= 0:
            raise ValueError("protein_limit must be positive")
        dataset = dataset[:protein_limit]

    return dataset


# Load the full dataset and join model inputs with experimental targets
def load_dataset(
    npz_dir,
    sasa_dir,
    excel_path,
    cutoff="C8",
    min_seq_sep=3,
    sheet_name=None,
    protein_limit=None,
    protein_ids=None,
):
    targets = load_excel_targets(excel_path, sheet_name=sheet_name)
    npz_files = sorted(glob.glob(os.path.join(npz_dir, "*.npz")))
    if not npz_files:
        raise ValueError(f"No .npz files found in folder: {npz_dir}")

    dataset, skipped = [], []
    for npz_path in npz_files:
        protein_id = extract_protein_id_from_npz_path(npz_path)
        if protein_id not in targets:
            skipped.append((protein_id, "not found in Excel"))
            continue

        sasa_path = get_sasa_path_for_protein(sasa_dir, protein_id)
        if not os.path.exists(sasa_path):
            skipped.append((protein_id, f"SASA file not found: {sasa_path}"))
            continue

        try:
            protein_data = load_single_protein_inputs(
                npz_path=npz_path,
                sasa_path=sasa_path,
                cutoff=cutoff,
                min_seq_sep=min_seq_sep,
            )
            protein_data["protein_id"] = protein_id
            protein_data["target_dG"] = targets[protein_id]
            dataset.append(protein_data)
        except Exception as e:
            skipped.append((protein_id, str(e)))

    dataset = select_dataset_subset(dataset, protein_limit=protein_limit, protein_ids=protein_ids)

    if skipped:
        print("\nSkipped proteins:")
        for pid, reason in skipped[:20]:
            print(f"  {pid}: {reason}")
        if len(skipped) > 20:
            print(f"  ... and {len(skipped) - 20} more")

    if not dataset:
        raise ValueError("No valid proteins loaded into dataset.")

    preview = ", ".join(item["protein_id"] for item in dataset[:10])
    suffix = "..." if len(dataset) > 10 else ""
    print(f"Loaded proteins for cutoff={cutoff}, min_seq_sep={min_seq_sep}: {len(dataset)}")
    print(f"Proteins used: {preview}{suffix}")
    return dataset


# Build segment weights for the WSME transfer-matrix calculation
def build_w_from_coefficients(contact_matrix, vdw_matrix, elec_matrix, entropy_vector, a1, b1, a2, b2, R, T):
    N = contact_matrix.shape[0]
    w = [[0.0] * (N + 2) for _ in range(N + 1)]

    entropy_prefix = np.zeros(N + 1, dtype=np.float64)
    entropy_prefix[1:] = np.cumsum(entropy_vector)

    pair_energy = contact_matrix * np.clip(a1 * vdw_matrix + b1 + a2 * elec_matrix + b2, -3.0, 1.0)
    pair_energy_prefix = build_2d_prefix_sum(pair_energy)

    for j in range(1, N + 1):
        for i in range(1, j + 1):
            start, end = i - 1, j - 1
            G_contacts = 0.5 * rect_sum(pair_energy_prefix, start, end, start, end)
            G_entropy = -T * (entropy_prefix[end + 1] - entropy_prefix[start])
            x = np.clip(-(G_contacts + G_entropy) / (R * T), -700.0, 700.0)
            w[j][i] = float(np.exp(x))
        w[j][j + 1] = 1.0

    w[0][1] = 1.0
    return w


# Transfer-matrix recursion over the number of ordered residues
def wsme_transfer_matrix_Zq(N, w):
    max_q = N
    vec_prev = np.zeros((N + 2, max_q + 1), dtype=np.float64)
    vec_prev[0, 0] = 1.0
    log_scale_total = 0.0

    for j in range(N, -1, -1):
        vec = np.zeros((j + 1, max_q + 1), dtype=np.float64)
        for k in range(1, j + 2):
            vec[k - 1, :] += vec_prev[k, :]
        from0 = vec_prev[0, :].copy()

        for k_target in range(0, j + 1):
            weight = w[j][j + 1 - k_target]
            if weight == 0.0:
                continue
            if k_target == 0:
                vec[k_target, :] += weight * from0
            else:
                vec[k_target, k_target:] += weight * from0[:max_q + 1 - k_target]

        max_val = np.max(vec)
        if max_val > 0 and np.isfinite(max_val):
            vec /= max_val
            log_scale_total += np.log(max_val)
        vec_prev = vec

    return vec_prev[0, :].copy(), log_scale_total


def compute_free_energy(Z_q, log_scale_total, R, T):
    F_q = np.full_like(Z_q, np.inf, dtype=np.float64)
    mask = (Z_q > 0) & np.isfinite(Z_q)
    if np.any(mask):
        F_q[mask] = -R * T * (np.log(Z_q[mask]) + log_scale_total)
    return F_q


# Smooth the profile before locating unfolded and folded minima
def smooth_curve(y, window=7):
    y = np.asarray(y, dtype=np.float64)
    if window <= 1 or len(y) < window:
        return y.copy()
    if window % 2 == 0:
        window += 1
    kernel = np.ones(window, dtype=np.float64) / window
    pad = window // 2
    y_pad = np.pad(y, pad_width=pad, mode="edge")
    return np.convolve(y_pad, kernel, mode="valid")


# Extract two-state metrics from the free-energy profile
def find_two_state_metrics(F_q_shifted, smooth_window=5):
    F_q_shifted = np.asarray(F_q_shifted, dtype=np.float64)
    N = len(F_q_shifted)
    split = N // 2

    if split < 2 or N - split < 2:
        raise ValueError("Profile too short.")

    F_smooth = smooth_curve(F_q_shifted, window=smooth_window)

    q_unf = int(np.argmin(F_smooth[:split]))
    q_fold = int(np.argmin(F_smooth[split:]) + split)

    if q_fold <= q_unf + 2:
        q_fold = min(N - 1, q_unf + 3)
    if q_fold <= q_unf + 1:
        raise ValueError("Could not separate unfolded and folded minima.")

    middle = F_smooth[q_unf + 1:q_fold]
    if len(middle) == 0:
        raise ValueError("No internal points for transition state.")

    q_ts = int(np.argmax(middle) + q_unf + 1)

    F_unf = float(F_q_shifted[q_unf])
    F_ts = float(F_q_shifted[q_ts])
    F_fold = float(F_q_shifted[q_fold])

    if not np.all(np.isfinite([F_unf, F_ts, F_fold])):
        raise ValueError("Non-finite two-state metrics")

    barrier = float(max(0.0, F_ts - F_unf))
    warning_parts = []

    if F_ts <= F_unf:
        warning_parts.append("barrier_below_unfolded")
    if q_ts <= q_unf or q_ts >= q_fold:
        warning_parts.append("ts_not_strictly_internal")

    return {
        "q_unf": int(q_unf),
        "q_ts": int(q_ts),
        "q_fold": int(q_fold),
        "F_unf": float(F_unf),
        "F_ts": float(F_ts),
        "F_fold": float(F_fold),
        "E_activation": barrier,
        "dG_folding": float((F_fold - F_unf) * (-1)),
        "barrier_warning": ";".join(warning_parts),
    }


# run the full thermodynamic model for one protein
def target_function(contact_matrix, vdw_matrix, elec_matrix, entropy_vector, a1, b1, a2, b2, R=R, T=TEMP, smooth_window=5):
    N = contact_matrix.shape[0]
    w = build_w_from_coefficients(contact_matrix, vdw_matrix, elec_matrix, entropy_vector, a1, b1, a2, b2, R, T)
    Z_q, log_scale_total = wsme_transfer_matrix_Zq(N, w)
    F_q = compute_free_energy(Z_q, log_scale_total, R, T)

    finite_mask = np.isfinite(F_q)
    if not np.any(finite_mask):
        raise ValueError("All F_q values are non-finite")

    F_q_shifted = F_q - np.min(F_q[finite_mask])
    metrics = find_two_state_metrics(F_q_shifted, smooth_window=smooth_window)
    return {
        "E_activation": metrics["E_activation"],
        "dG_folding": metrics["dG_folding"],
        "q_unf": metrics["q_unf"],
        "q_ts": metrics["q_ts"],
        "q_fold": metrics["q_fold"],
        "barrier_warning": metrics["barrier_warning"],
        "F_q_shifted": F_q_shifted,
    }


def build_entropy_for_item(item, entropy_aa_params):
    entropy_vector = build_entropy_vector_from_sasa_per_aa(
        res_name3=item["res_name3"],
        sasa_norm=item["sasa_norm"],
        entropy_aa_params=entropy_aa_params,
    )
    return -np.abs(entropy_vector)


# keep entropy parameters in a physically reasonable range
def entropy_regularization(
    entropy_aa_params,
    entropy_l2=0.0,
    penalize_positive_entropy=False,
    positive_entropy_penalty=100.0,
    entropy_abs_limit=0.05,
    entropy_abs_penalty=1000.0,
):
    penalty = 0.0

    if entropy_l2 > 0.0:
        penalty += entropy_l2 * float(np.sum(entropy_aa_params ** 2))

    if penalize_positive_entropy:
        pos = np.clip(entropy_aa_params, 0.0, None)
        penalty += positive_entropy_penalty * float(np.sum(pos ** 2))

    if entropy_abs_limit is not None and entropy_abs_limit > 0.0:
        excess = np.clip(np.abs(entropy_aa_params) - entropy_abs_limit, 0.0, None)
        penalty += entropy_abs_penalty * float(np.sum(excess ** 2))

    return penalty


# objective for global fitting across all selected proteins
def objective_function_global(
    params,
    dataset,
    smooth_window=5,
    entropy_l2=0.0,
    penalize_positive_entropy=False,
    positive_entropy_penalty=100.0,
    entropy_abs_limit=0.05,
    entropy_abs_penalty=1000.0,
):
    call_counter["n"] += 1
    a1, b1, a2, b2 = params[:4]
    entropy_aa_params = params[4:4 + len(AA_ORDER)]

    penalty_params = entropy_regularization(
        entropy_aa_params=entropy_aa_params,
        entropy_l2=entropy_l2,
        penalize_positive_entropy=penalize_positive_entropy,
        positive_entropy_penalty=positive_entropy_penalty,
        entropy_abs_limit=entropy_abs_limit,
        entropy_abs_penalty=entropy_abs_penalty,
    )

    total_loss = 0.0
    n_valid = 0
    invalid_ids = []
    invalid_details = []

    for item in dataset:
        try:
            result = target_function(
                contact_matrix=item["contact_matrix"],
                vdw_matrix=item["vdw_matrix"],
                elec_matrix=item["elec_matrix"],
                entropy_vector=build_entropy_for_item(item, entropy_aa_params),
                a1=a1,
                b1=b1,
                a2=a2,
                b2=b2,
                R=R,
                T=TEMP,
                smooth_window=smooth_window,
            )
            dG_pred = result["dG_folding"]
            if not np.isfinite(dG_pred):
                invalid_ids.append(item["protein_id"])
                invalid_details.append((item["protein_id"], "predicted dG is non-finite"))
                total_loss += INVALID_PROTEIN_PENALTY
                continue

            total_loss += (dG_pred - item["target_dG"]) ** 2
            n_valid += 1
        except Exception as e:
            invalid_ids.append(item["protein_id"])
            invalid_details.append((item["protein_id"], str(e)))
            total_loss += INVALID_PROTEIN_PENALTY

    if n_valid == 0:
        return 1e12

    mean_loss = total_loss / n_valid + penalty_params

    if mean_loss < best_tracker["best_loss"]:
        best_tracker["best_loss"] = mean_loss
        best_tracker["best_params"] = np.array(params, dtype=np.float64).copy()
        best_tracker["invalid_ids"] = invalid_ids[:]
        best_tracker["invalid_details"] = invalid_details[:]

    if call_counter["n"] % 10 == 0:
        print(
            f"[pid {os.getpid()}] call {call_counter['n']}: "
            f"loss={mean_loss:.5f}, best={best_tracker['best_loss']:.5f}, "
            f"a1={a1:.5f}, b1={b1:.5f}, a2={a2:.5f}, b2={b2:.5f}, "
            f"n_valid={n_valid}, n_invalid={len(invalid_ids)}"
        )
    return float(mean_loss)


# evaluate the fitted model protein by protein
def evaluate_dataset(dataset, params, smooth_window=5):
    a1, b1, a2, b2 = params[:4]
    entropy_aa_params = params[4:4 + len(AA_ORDER)]
    rows = []

    for item in dataset:
        try:
            result = target_function(
                contact_matrix=item["contact_matrix"],
                vdw_matrix=item["vdw_matrix"],
                elec_matrix=item["elec_matrix"],
                entropy_vector=build_entropy_for_item(item, entropy_aa_params),
                a1=a1,
                b1=b1,
                a2=a2,
                b2=b2,
                R=R,
                T=TEMP,
                smooth_window=smooth_window,
            )
            dG_pred = result["dG_folding"]
            err = dG_pred - item["target_dG"]

            rows.append({
                "protein_id": item["protein_id"],
                "target_dG": item["target_dG"],
                "predicted_dG": dG_pred,
                "error": err,
                "sq_error": err ** 2,
                "E_activation": result["E_activation"],
                "q_unf": result["q_unf"],
                "q_ts": result["q_ts"],
                "q_fold": result["q_fold"],
                "barrier_warning": result["barrier_warning"],
                "status": "ok",
                "failure_reason": "",
            })
        except Exception as e:
            rows.append({
                "protein_id": item["protein_id"],
                "target_dG": item["target_dG"],
                "predicted_dG": np.nan,
                "error": np.nan,
                "sq_error": np.nan,
                "E_activation": np.nan,
                "q_unf": np.nan,
                "q_ts": np.nan,
                "q_fold": np.nan,
                "barrier_warning": "",
                "status": "failed",
                "failure_reason": str(e),
            })
    return rows


def compute_metrics_from_rows(rows):
    sq_errors = [r["sq_error"] for r in rows if np.isfinite(r["sq_error"])]
    abs_errors = [abs(r["error"]) for r in rows if np.isfinite(r["error"])]
    return {
        "n_valid": len(sq_errors),
        "rmse": np.sqrt(np.mean(sq_errors)) if sq_errors else np.nan,
        "mae": np.mean(abs_errors) if abs_errors else np.nan,
    }


def save_csv(path, rows, fieldnames):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def save_results_csv(path, rows):
    save_csv(path, rows, [
        "protein_id", "target_dG", "predicted_dG", "error", "sq_error",
        "E_activation", "q_unf", "q_ts", "q_fold",
        "barrier_warning",
        "status", "failure_reason"
    ])


def write_restart_params_file(path, params):
    with open(path, "w", encoding="utf-8") as f:
        f.write(",".join(f"{x:.12g}" for x in params))
        f.write("\n")


def parse_restart_params_string(s):
    values = [to_float_safe(x) for x in s.replace("\n", ",").split(",")]
    values = [x for x in values if x is not None]
    expected = 4 + len(AA_ORDER)
    if len(values) != expected:
        raise ValueError(f"Expected {expected} restart parameters, got {len(values)}")
    return np.array(values, dtype=np.float64)


def load_restart_params(restart_params=None, restart_params_file=None):
    provided = int(restart_params is not None) + int(restart_params_file is not None)
    if provided > 1:
        raise ValueError("Use only one of --restart_params or --restart_params_file")
    if restart_params is not None:
        return parse_restart_params_string(restart_params)
    if restart_params_file is not None:
        with open(restart_params_file, "r", encoding="utf-8") as f:
            return parse_restart_params_string(f.read())
    return None


# Run optimization experiment
def run_single_experiment(
    npz_dir,
    sasa_dir,
    excel_path,
    cutoff,
    min_seq_sep,
    sheet_name=None,
    maxfev=150,
    smooth_window=5,
    entropy_init=-0.01,
    entropy_l2=0.1,
    penalize_positive_entropy=False,
    positive_entropy_penalty=100.0,
    entropy_abs_limit=0.05,
    entropy_abs_penalty=1000.0,
    restart_params=None,
    protein_limit=None,
    protein_ids=None,
):
    dataset = load_dataset(
        npz_dir=npz_dir,
        sasa_dir=sasa_dir,
        excel_path=excel_path,
        cutoff=cutoff,
        min_seq_sep=min_seq_sep,
        sheet_name=sheet_name,
        protein_limit=protein_limit,
        protein_ids=protein_ids,
    )

    default_x0 = np.concatenate([
        np.array([0.1, 0.1, 0.1, 0.1], dtype=np.float64),
        get_initial_entropy_params_per_aa(default_value=entropy_init),
    ])
    x0 = np.array(restart_params, dtype=np.float64).copy() if restart_params is not None else default_x0

    if len(x0) != 4 + len(AA_ORDER):
        raise ValueError(f"x0 has length {len(x0)}, expected {4 + len(AA_ORDER)}")

    print(
        f"\n[pid {os.getpid()}] cutoff={cutoff}, min_seq_sep={min_seq_sep}, "
        f"n_params={len(x0)}, maxfev={maxfev}, smooth_window={smooth_window}, "
        f"entropy_init={entropy_init}, entropy_l2={entropy_l2}, "
        f"restart_used={restart_params is not None}, protein_limit={protein_limit}, protein_ids={protein_ids}"
    )

    call_counter["n"] = 0
    best_tracker["best_loss"] = np.inf
    best_tracker["best_params"] = None
    best_tracker["invalid_ids"] = []
    best_tracker["invalid_details"] = []

    result_opt = minimize(
        fun=objective_function_global,
        x0=x0,
        args=(
            dataset,
            smooth_window,
            entropy_l2,
            penalize_positive_entropy,
            positive_entropy_penalty,
            entropy_abs_limit,
            entropy_abs_penalty,
        ),
        method="Powell",
        options={"maxfev": maxfev, "disp": False, "xtol": 1e-3, "ftol": 1e-3},
    )

    final_params = best_tracker["best_params"] if best_tracker["best_params"] is not None else result_opt.x
    rows = evaluate_dataset(dataset, final_params, smooth_window=smooth_window)
    metrics = compute_metrics_from_rows(rows)

    return {
        "cutoff": cutoff,
        "min_seq_sep": min_seq_sep,
        "params": final_params,
        "final_loss": best_tracker["best_loss"] if np.isfinite(best_tracker["best_loss"]) else result_opt.fun,
        "rmse": metrics["rmse"],
        "mae": metrics["mae"],
        "n_valid": metrics["n_valid"],
        "success": result_opt.success,
        "message": str(result_opt.message),
        "rows": rows,
        "best_invalid_ids": best_tracker.get("invalid_ids", []),
        "best_invalid_details": best_tracker.get("invalid_details", []),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--npz_dir", required=True)
    parser.add_argument("--sasa_dir", required=True)
    parser.add_argument("--excel", required=True)
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--maxfev", type=int, default=150)
    parser.add_argument("--cutoff", default="C8", choices=["C4", "C6", "C8"])
    parser.add_argument("--min_seq_sep", type=int, default=3)
    parser.add_argument("--smooth_window", type=int, default=5)
    parser.add_argument("--entropy_init", type=float, default=-0.01)
    parser.add_argument("--entropy_l2", type=float, default=0.1)
    parser.add_argument("--penalize_positive_entropy", action="store_true")
    parser.add_argument("--positive_entropy_penalty", type=float, default=100.0)
    parser.add_argument("--entropy_abs_limit", type=float, default=0.05)
    parser.add_argument("--entropy_abs_penalty", type=float, default=1000.0)
    parser.add_argument("--restart_params", type=str, default=None)
    parser.add_argument("--restart_params_file", type=str, default=None)
    parser.add_argument("--protein_limit", type=int, default=None)
    parser.add_argument("--protein_ids", nargs="+", default=None)
    args = parser.parse_args()

    restart_params = load_restart_params(args.restart_params, args.restart_params_file)

    result = run_single_experiment(
        npz_dir=args.npz_dir,
        sasa_dir=args.sasa_dir,
        excel_path=args.excel,
        cutoff=args.cutoff,
        min_seq_sep=args.min_seq_sep,
        sheet_name=args.sheet,
        maxfev=args.maxfev,
        smooth_window=args.smooth_window,
        entropy_init=args.entropy_init,
        entropy_l2=args.entropy_l2,
        penalize_positive_entropy=args.penalize_positive_entropy,
        positive_entropy_penalty=args.positive_entropy_penalty,
        entropy_abs_limit=args.entropy_abs_limit,
        entropy_abs_penalty=args.entropy_abs_penalty,
        restart_params=restart_params,
        protein_limit=args.protein_limit,
        protein_ids=args.protein_ids,
    )

    params = result["params"]
    a1_opt, b1_opt, a2_opt, b2_opt = params[:4]
    entropy_aa_params_opt = params[4:4 + len(AA_ORDER)]

    print("\n=== GLOBAL SASA FIT RESULT (PER-AA ENTROPY, 1 - NORMALIZED SASA) ===")
    for k, v in [
        ("cutoff", result["cutoff"]),
        ("min_seq_sep", result["min_seq_sep"]),
        ("Final loss", result["final_loss"]),
        ("RMSE", result["rmse"]),
        ("MAE", result["mae"]),
        ("n_valid", result["n_valid"]),
        ("Optimization success", result["success"]),
        ("Message", result["message"]),
    ]:
        print(f"{k} = {v}")

    print("\n=== ENERGY PARAMETERS ===")
    print("a1 =", a1_opt)
    print("b1 =", b1_opt)
    print("a2 =", a2_opt)
    print("b2 =", b2_opt)

    print("\n=== PER-AA ENTROPY PARAMETERS ===")
    for aa, val in zip(AA_ORDER, entropy_aa_params_opt):
        print(f"{aa:>3s} = {val:.8f}")

    if result["best_invalid_details"]:
        print("\n=== INVALID PROTEINS AT BEST POINT ===")
        for pid, reason in result["best_invalid_details"]:
            print(f"{pid}: {reason}")
    else:
        print("\n=== INVALID PROTEINS AT BEST POINT ===")
        print("none")

    out_csv = f"global_sasa_fit_perAA_results_{args.cutoff}_sep{args.min_seq_sep}_maxfev{args.maxfev}.csv"
    save_results_csv(out_csv, result["rows"])
    print(f"\nSaved per-protein results to: {out_csv}")

    restart_out = f"restart_params_{args.cutoff}_sep{args.min_seq_sep}_maxfev{args.maxfev}.txt"
    write_restart_params_file(restart_out, params)
    print(f"Saved restart parameters to: {restart_out}")


if __name__ == "__main__":
    main()
